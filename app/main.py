from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, Request, Form, Depends, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, text
from app.database import engine, SessionLocal, get_db
from app import models
from app.models import (
    Client, Quote, Invoice, User, Service, PasswordResetToken, 
    QuoteItem, InvoiceItem, AuditLog, UserRole
)
from app.emailer import send_email
from app.pdf import generate_quote_pdf
from app.invoice_pdf import generate_invoice_pdf
import json
import os
import secrets
import re
from datetime import datetime, timedelta
from passlib.context import CryptContext
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from starlette.responses import StreamingResponse
from openpyxl.drawing.image import Image as XLImage
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# SECURITY CONFIG
EMAIL_REGEX = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
pwd_context = CryptContext(schemes=["argon2"], deprecated="auto")

def hash_password(password: str):
    return pwd_context.hash(password)

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def require_auth(request: Request, db: Session = Depends(get_db)):
    """Ensure user is logged in"""
    user_id = request.session.get("user_id")
    if not user_id:
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    
    user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
    if not user:
        request.session.clear()
        raise HTTPException(status_code=303, headers={"Location": "/login"})
    
    return user

def require_admin(current_user: User = Depends(require_auth)):
    """Ensure user is admin"""
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

def generate_csrf_token():
    return secrets.token_urlsafe(32)

def log_audit_action(db: Session, user_id: int, action: str, entity_type: str = None, 
                     entity_id: int = None, details: str = None, ip_address: str = None):
    log = AuditLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details,
        ip_address=ip_address
    )
    db.add(log)
    db.commit()

# APP SETUP
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Umvuzo Media Invoicing System - Secure")

# CSRF Token Middleware
@app.middleware("http")
async def add_csrf_token(request: Request, call_next):
    if "session" in request.scope:
        if "csrf_token" not in request.session:
            request.session["csrf_token"] = generate_csrf_token()
    response = await call_next(request)
    return response

# Security Middleware
SECRET_KEY = os.getenv("SESSION_SECRET", secrets.token_urlsafe(32))
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, max_age=3600)

class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        return response

app.add_middleware(SecurityHeadersMiddleware)

app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")

# =========================
# AUTH ROUTES
# =========================

@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if request.session.get("user_id"):
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse("login.html", {
        "request": request,
        "csrf_token": request.session.get("csrf_token", generate_csrf_token())
    })

@app.post("/login")
def login_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    csrf_token: str = Form(...),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    user = db.query(User).filter(User.username == username, User.is_active == True).first()
    
    if not user or not verify_password(password, user.password):
        request.session["flash"] = "Invalid credentials."
        return RedirectResponse("/login", status_code=303)
    
    user.last_login = datetime.utcnow()
    db.commit()
    
    request.session["user_id"] = user.id
    request.session["csrf_token"] = generate_csrf_token()
    
    return RedirectResponse("/dashboard", status_code=303)


@app.get("/", response_class=HTMLResponse)
def root_redirect(request: Request):
    """Redirect root URL to login page"""
    if request.session.get("user_id"):
        return RedirectResponse("/dashboard", status_code=302)
    return RedirectResponse("/login", status_code=302)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=302)

@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page(request: Request):
    return templates.TemplateResponse("forgot_password.html", {
        "request": request,
        "csrf_token": request.session.get("csrf_token", generate_csrf_token())
    })

@app.post("/forgot-password")
def forgot_password(
    request: Request, 
    username: str = Form(...),
    csrf_token: str = Form(...),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    user = db.query(User).filter(User.username == username).first()
    
    flash_message = "If that email exists, a reset link has been sent."
    
    if user:
        db.query(PasswordResetToken).filter(
            PasswordResetToken.user_id == user.id,
            PasswordResetToken.used == False
        ).update({"used": True})
        
        token = secrets.token_urlsafe(32)
        reset = PasswordResetToken(
            user_id=user.id,
            token=token,
            expires_at=datetime.utcnow() + timedelta(minutes=30)
        )
        db.add(reset)
        db.commit()
        
        BASE_URL = os.getenv("BASE_URL", "http://127.0.0.1:8000")
        reset_link = f"{BASE_URL}/reset-password/{token}"
        
        try:
            send_email(
                to_email=user.username,
                subject="Password Reset - Umvuzo Invoicing",
                body=f"""Hello,\n\nClick to reset:\n{reset_link}\n\nExpires in 30 minutes.""",
                pdf_path=None
            )
        except:
            pass
    
    request.session["flash"] = flash_message
    return RedirectResponse("/login", status_code=303)

@app.get("/reset-password/{token}", response_class=HTMLResponse)
def reset_password_page(request: Request, token: str, db: Session = Depends(get_db)):
    reset = db.query(PasswordResetToken).filter(
        PasswordResetToken.token == token,
        PasswordResetToken.used == False
    ).first()
    
    if not reset or reset.expires_at < datetime.utcnow():
        request.session["flash"] = "Invalid or expired link."
        return RedirectResponse("/login", status_code=303)
    
    return templates.TemplateResponse("reset_password.html", {
        "request": request,
        "token": token,
        "csrf_token": request.session.get("csrf_token", generate_csrf_token())
    })

@app.post("/reset-password/{token}")
def reset_password(
    request: Request,
    token: str,
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    csrf_token: str = Form(...),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    reset = db.query(PasswordResetToken).filter(
        PasswordResetToken.token == token,
        PasswordResetToken.used == False
    ).first()
    
    if not reset or reset.expires_at < datetime.utcnow():
        request.session["flash"] = "Invalid link."
        return RedirectResponse("/login", status_code=303)
    
    if new_password != confirm_password:
        request.session["flash"] = "Passwords don't match."
        return RedirectResponse(f"/reset-password/{token}", status_code=303)
    
    if len(new_password) < 8:
        request.session["flash"] = "Password too short (min 8)."
        return RedirectResponse(f"/reset-password/{token}", status_code=303)
    
    user = db.query(User).get(reset.user_id)
    user.password = hash_password(new_password)
    reset.used = True
    db.commit()
    
    log_audit_action(db, user.id, "password_reset", ip_address=request.client.host)
    
    request.session["flash"] = "Password reset!"
    return RedirectResponse("/login", status_code=303)

# =========================
# DASHBOARD
# =========================

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    if current_user.role == UserRole.ADMIN:
        data = {
            "total_quotes": db.query(Quote).count(),
            "approved_quotes": db.query(Quote).filter(Quote.status == "Approved").count(),
            "converted_quotes": db.query(Quote).filter(Quote.converted == True).count(),
            "total_invoices": db.query(Invoice).count(),
            "paid_invoices": db.query(Invoice).filter(Invoice.paid == True).count(),
            "unpaid_invoices": db.query(Invoice).filter(Invoice.paid == False).count(),
        }
    else:
        data = {
            "total_quotes": db.query(Quote).filter(Quote.created_by_id == current_user.id).count(),
            "approved_quotes": db.query(Quote).filter(Quote.created_by_id == current_user.id, Quote.status == "Approved").count(),
            "converted_quotes": db.query(Quote).filter(Quote.created_by_id == current_user.id, Quote.converted == True).count(),
            "total_invoices": db.query(Invoice).filter(Invoice.created_by_id == current_user.id).count(),
            "paid_invoices": db.query(Invoice).filter(Invoice.created_by_id == current_user.id, Invoice.paid == True).count(),
            "unpaid_invoices": db.query(Invoice).filter(Invoice.created_by_id == current_user.id, Invoice.paid == False).count(),
        }
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request, 
        "current_user": current_user,
        **data
    })

# =========================
# USER MANAGEMENT (Admin Only)
# =========================

@app.get("/users/create", response_class=HTMLResponse)
def create_user_page(request: Request, current_user: User = Depends(require_admin)):
    return templates.TemplateResponse("create_user.html", {
        "request": request,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/users/create")
def create_user(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form("user"),
    csrf_token: str = Form(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    if not re.match(EMAIL_REGEX, username):
        request.session["flash"] = "Username must be a valid email."
        return RedirectResponse("/users/create", status_code=303)
    
    if len(password) < 8:
        request.session["flash"] = "Password must be at least 8 characters."
        return RedirectResponse("/users/create", status_code=303)
    
    existing = db.query(User).filter(User.username == username).first()
    if existing:
        request.session["flash"] = "User already exists."
        return RedirectResponse("/users/create", status_code=303)
    
    user = User(
        username=username,
        password=hash_password(password),
        role=UserRole.ADMIN if role == "admin" else UserRole.USER
    )
    db.add(user)
    db.commit()
    
    log_audit_action(db, current_user.id, "user_created", "user", user.id, 
                    f"Created {username}", request.client.host)
    
    request.session["flash"] = "User created!"
    return RedirectResponse("/dashboard", status_code=303)

@app.get("/change-password", response_class=HTMLResponse)
def change_password_page(request: Request, current_user: User = Depends(require_auth)):
    return templates.TemplateResponse("change_password.html", {
        "request": request,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/change-password")
def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    csrf_token: str = Form(...),
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    if not verify_password(current_password, current_user.password):
        request.session["flash"] = "Current password is wrong."
        return RedirectResponse("/change-password", status_code=303)
    
    if new_password != confirm_password:
        request.session["flash"] = "Passwords don't match."
        return RedirectResponse("/change-password", status_code=303)
    
    if len(new_password) < 8:
        request.session["flash"] = "Password too short."
        return RedirectResponse("/change-password", status_code=303)
    
    current_user.password = hash_password(new_password)
    db.commit()
    
    log_audit_action(db, current_user.id, "password_changed", ip_address=request.client.host)
    
    request.session["flash"] = "Password updated!"
    return RedirectResponse("/dashboard", status_code=303)

# =========================
# CLIENTS
# =========================

@app.get("/clients-page", response_class=HTMLResponse)
def clients_page(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    clients = db.query(Client).all()  # Changed: Everyone sees all clients
    
    return templates.TemplateResponse("clients.html", {
        "request": request, 
        "clients": clients,
        "current_user": current_user
    })

@app.get("/clients/create", response_class=HTMLResponse)
def create_client_form(request: Request, current_user: User = Depends(require_auth)):
    return templates.TemplateResponse("create_client.html", {
        "request": request,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/clients/create")
def create_client(
    request: Request,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    address: str = Form(None),
    billing_name: str = Form(None),
    billing_email: str = Form(None),
    billing_address: str = Form(None),
    vat_number: str = Form(None),
    tax_number: str = Form(None),
    payment_terms: str = Form(None),
    csrf_token: str = Form(...),
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
      
    try:
        clean_name = re.sub(r'[^A-Za-z]', '', name).upper()
        if len(clean_name) < 3:
            clean_name = (clean_name + "XXX")[:3]
        
        base_code = clean_name[:3]
        existing = db.query(Client).filter(Client.client_code.like(f"{base_code}%")).all()
        
        if existing:
            numbers = []
            for c in existing:
                try:
                    num = int(c.client_code[3:])
                    numbers.append(num)
                except:
                    continue
            next_num = max(numbers) + 1 if numbers else 1
        else:
            next_num = 1
        
        client_code = f"{base_code}{next_num:03d}"
        
        client = Client(
            name=name, email=email, phone=phone, address=address,
            client_code=client_code, billing_name=billing_name,
            billing_email=billing_email, billing_address=billing_address,
            vat_number=vat_number, tax_number=tax_number,
            payment_terms=payment_terms, created_by_id=current_user.id
        )
        db.add(client)
        db.commit()
        
        log_audit_action(db, current_user.id, "client_created", "client", client.id, 
                        f"Created {name} ({client_code})", request.client.host)
        
        request.session["flash"] = f"Client created: {client_code}"
        return RedirectResponse("/clients-page", status_code=303)
        
    except Exception as e:
        db.rollback()
        request.session["flash"] = "Error creating client."
        return RedirectResponse("/clients/create", status_code=303)

@app.get("/clients/{client_id}/edit", response_class=HTMLResponse)
def edit_client_page(client_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        return RedirectResponse("/clients-page", status_code=303)
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return templates.TemplateResponse("edit_client.html", {
        "request": request, "client": client,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/clients/{client_id}/edit")
def update_client(
    client_id: int, request: Request, name: str = Form(...),
    email: str = Form(...), phone: str = Form(...), address: str = Form(None),
    billing_name: str = Form(None), billing_email: str = Form(None),
    billing_address: str = Form(None), vat_number: str = Form(None),
    tax_number: str = Form(None), payment_terms: str = Form(None),
    csrf_token: str = Form(...), current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        return RedirectResponse("/clients-page", status_code=303)
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client.name = name
    client.email = email
    client.phone = phone
    client.address = address
    client.billing_name = billing_name
    client.billing_email = billing_email
    client.billing_address = billing_address
    client.vat_number = vat_number
    client.tax_number = tax_number
    client.payment_terms = payment_terms
    
    db.commit()
    log_audit_action(db, current_user.id, "client_updated", "client", client.id, 
                    f"Updated {name}", request.client.host)
    
    return RedirectResponse("/clients-page", status_code=303)

@app.get("/clients/{client_id}", response_class=HTMLResponse)
def client_history(client_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    client = db.get(Client, client_id)
    if not client:
        return RedirectResponse("/clients-page")
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user.role == UserRole.ADMIN:
        quotes = db.query(Quote).filter(Quote.client_id == client_id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id).all()
    else:
        quotes = db.query(Quote).filter(Quote.client_id == client_id, Quote.created_by_id == current_user.id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id, Invoice.created_by_id == current_user.id).all()
    
    return templates.TemplateResponse("client_history.html", {
        "request": request, "client": client, "quotes": quotes,
        "invoices": invoices, "current_user": current_user
    })

# =========================
# QUOTES
# =========================

@app.get("/quotes-page", response_class=HTMLResponse)
def quotes_page(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    if current_user.role == UserRole.ADMIN:
        quotes = db.query(Quote).options(joinedload(Quote.items), joinedload(Quote.client)).all()
    else:
        quotes = db.query(Quote).filter(Quote.created_by_id == current_user.id).options(
            joinedload(Quote.items), joinedload(Quote.client)).all()
    
    return templates.TemplateResponse("quotes.html", {
        "request": request, "quotes": quotes, "current_user": current_user
    })

@app.get("/quotes/create", response_class=HTMLResponse)
def create_quote_form(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    clients = db.query(Client).all()  # Changed: Everyone sees all clients
    services = db.query(Service).filter(Service.is_active == True).order_by(Service.name).all()
    
    return templates.TemplateResponse("create_quote.html", {
        "request": request, "clients": clients, "services": services,
        "csrf_token": request.session.get("csrf_token"), "current_user": current_user
    })

@app.post("/quotes/create")
def create_quote(
    request: Request, client_id: int = Form(...), items_data: str = Form(...),
    csrf_token: str = Form(...), current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client or (current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id):
        raise HTTPException(status_code=403, detail="Invalid client")
    
    try:
        items = json.loads(items_data)
        if not items:
            request.session["flash"] = "Add at least one item."
            return RedirectResponse("/quotes/create", status_code=303)
    except:
        request.session["flash"] = "Invalid data."
        return RedirectResponse("/quotes/create", status_code=303)
    
    try:
        last = db.query(func.max(Quote.quote_number)).filter(Quote.client_id == client_id).scalar()
        next_num = 1 if not last else last + 1
        
        quote = Quote(
            quote_number=next_num, client_id=client_id, total=0,
            status="Draft", converted=False, created_by_id=current_user.id
        )
        db.add(quote)
        db.flush()
        
        total = 0
        for item in items:
            line = float(item["unit_cost"]) * float(item["quantity"])
            total += line
            db.add(QuoteItem(
                quote_id=quote.id, description=item["description"],
                unit_cost=float(item["unit_cost"]), quantity=float(item["quantity"])
            ))
        
        quote.total = total
        db.commit()
        
        log_audit_action(db, current_user.id, "quote_created", "quote", quote.id, 
                        f"Q-{next_num:04d} for {client.name}", request.client.host)
        return RedirectResponse("/quotes-page", status_code=303)
    except:
        db.rollback()
        request.session["flash"] = "Error creating quote."
        return RedirectResponse("/quotes/create", status_code=303)

@app.get("/quotes/{quote_id}/convert")
def convert_quote(quote_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).with_for_update().first()
    
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if quote.converted:
        request.session["flash"] = "Already converted."
        return RedirectResponse("/quotes-page", status_code=303)
    
    if quote.status != "Approved":
        request.session["flash"] = "Quote must be approved first."
        return RedirectResponse("/quotes-page", status_code=303)
    
    try:
        last_inv = db.query(func.max(Invoice.invoice_number)).filter(Invoice.client_id == quote.client_id).scalar()
        next_inv = 1 if not last_inv else last_inv + 1
        
        invoice = Invoice(
            invoice_number=next_inv, client_id=quote.client_id,
            total=0, paid=False, created_by_id=current_user.id
        )
        db.add(invoice)
        db.flush()
        
        quote_items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).all()
        total = 0
        for item in quote_items:
            line = item.unit_cost * item.quantity
            total += line
            db.add(InvoiceItem(
                invoice_id=invoice.id, description=item.description,
                unit_cost=item.unit_cost, quantity=item.quantity
            ))
        
        invoice.total = total
        quote.converted = True
        db.commit()
        
        log_audit_action(db, current_user.id, "quote_converted", "invoice", invoice.id, 
                        f"Q-{quote.quote_number:04d} to INV-{next_inv:04d}", request.client.host)
        return RedirectResponse("/invoices-page", status_code=303)
    except:
        db.rollback()
        request.session["flash"] = "Error converting."
        return RedirectResponse("/quotes-page", status_code=303)

@app.get("/quotes/{quote_id}/pdf")
def quote_pdf(quote_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.get(Quote, quote_id)
    if not quote:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client = db.get(Client, quote.client_id)
    items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).all()
    filename = f"quote_{quote.id}.pdf"
    generate_quote_pdf(quote, client, items, filename)
    return FileResponse(filename, media_type="application/pdf", filename=filename)

@app.get("/quotes/{quote_id}/email")
def email_quote(quote_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        return RedirectResponse("/quotes-page", status_code=303)
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).all()
        filename = f"quote_{quote.quote_number}.pdf"
        generate_quote_pdf(quote, quote.client, items, filename)
        
        send_email(
            quote.client.email,
            f"Quote #{quote.quote_number:04d}",
            f"Dear {quote.client.name},\n\nPlease find your quote attached.\n\nTotal: R {quote.total:.2f}",
            filename
        )
        
        request.session["flash"] = "Quote emailed!"
        log_audit_action(db, current_user.id, "quote_emailed", "quote", quote.id, 
                        f"To {quote.client.email}", request.client.host)
    except Exception as e:
        request.session["flash"] = f"Email failed: {str(e)}"
    
    return RedirectResponse("/quotes-page", status_code=303)

@app.get("/quotes/{quote_id}/approved")
def approve_quote(quote_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    quote.status = "Approved"
    db.commit()
    log_audit_action(db, current_user.id, "quote_approved", "quote", quote.id, ip_address=request.client.host)
    return RedirectResponse("/quotes-page", status_code=303)

@app.get("/quotes/{quote_id}/sent")
def mark_sent(quote_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    quote.status = "Sent"
    db.commit()
    log_audit_action(db, current_user.id, "quote_sent", "quote", quote.id, ip_address=request.client.host)
    return RedirectResponse("/quotes-page", status_code=303)

# =========================
# INVOICES
# =========================

@app.get("/invoices-page", response_class=HTMLResponse)
def invoices_page(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    if current_user.role == UserRole.ADMIN:
        invoices = db.query(Invoice).options(joinedload(Invoice.items), joinedload(Invoice.client)).all()
    else:
        invoices = db.query(Invoice).filter(Invoice.created_by_id == current_user.id).options(
            joinedload(Invoice.items), joinedload(Invoice.client)).all()
    
    return templates.TemplateResponse("invoices.html", {
        "request": request, "invoices": invoices, "current_user": current_user
    })

@app.get("/invoices/{invoice_id}/paid")
def mark_paid(invoice_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and invoice.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if not invoice.paid:
        invoice.paid = True
        invoice.paid_at = datetime.utcnow()
        invoice.marked_paid_by_id = current_user.id
        db.commit()
        log_audit_action(db, current_user.id, "invoice_marked_paid", "invoice", invoice.id, ip_address=request.client.host)
    
    return RedirectResponse("/invoices-page", status_code=303)

@app.get("/invoices/{invoice_id}/pdf")
def invoice_pdf(invoice_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and invoice.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client = db.get(Client, invoice.client_id)
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice.id).all()
    filename = f"invoice_{invoice.id}.pdf"
    generate_invoice_pdf(invoice, client, items, filename, client.client_code)
    return FileResponse(filename, media_type="application/pdf", filename=filename)

@app.get("/invoices/{invoice_id}/email")
def email_invoice(request: Request, invoice_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    invoice = db.get(Invoice, invoice_id)
    if not invoice:
        raise HTTPException(status_code=404, detail="Not found")
    
    if current_user.role != UserRole.ADMIN and invoice.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client = db.get(Client, invoice.client_id)
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice.id).all()
    filename = f"invoice_{invoice.id}.pdf"
    generate_invoice_pdf(invoice, client, items, filename, client.client_code)
    
    try:
        send_email(
            to_email=client.email,
            subject=f"Invoice {client.client_code}-INV-{invoice.invoice_number:04d}",
            body=f"Dear {client.name},\n\nPlease find invoice attached.\n\nTotal: R {invoice.total:.2f}\n{'PAID' if invoice.paid else 'PENDING'}",
            pdf_path=filename
        )
        request.session["flash"] = "Invoice emailed!"
        log_audit_action(db, current_user.id, "invoice_emailed", "invoice", invoice.id, 
                        f"To {client.email}", request.client.host)
    except Exception as e:
        request.session["flash"] = f"Email failed: {str(e)}"
    
    return RedirectResponse("/invoices-page", status_code=303)

# =========================
# SERVICES
# =========================

@app.get("/services", response_class=HTMLResponse)
def services_page(request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    services = db.query(Service).filter(Service.is_active == True).order_by(Service.category, Service.name).all()
    return templates.TemplateResponse("services.html", {
        "request": request, "services": services,
        "current_user": current_user, "csrf_token": request.session.get("csrf_token")
    })

@app.post("/services/create")
def create_service(
    request: Request, name: str = Form(...), description: str = Form(...),
    price: float = Form(...), category: str = Form(...),
    csrf_token: str = Form(...), current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    service = Service(name=name, description=description, price=price, category=category, is_active=True)
    db.add(service)
    db.commit()
    log_audit_action(db, current_user.id, "service_created", "service", service.id, name, request.client.host)
    return RedirectResponse("/services", status_code=303)

@app.get("/services/{service_id}/edit", response_class=HTMLResponse)
def edit_service_page(service_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    service = db.query(Service).filter(Service.id == service_id, Service.is_active == True).first()
    if not service:
        return RedirectResponse("/services", status_code=303)
    return templates.TemplateResponse("edit_service.html", {
        "request": request, "service": service,
        "current_user": current_user, "csrf_token": request.session.get("csrf_token")
    })

@app.post("/services/{service_id}/edit")
def update_service(
    service_id: int, request: Request, name: str = Form(...),
    description: str = Form(...), price: float = Form(...), category: str = Form(...),
    csrf_token: str = Form(...), current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    service = db.query(Service).filter(Service.id == service_id).first()
    if not service:
        return RedirectResponse("/services", status_code=303)
    
    service.name = name
    service.description = description
    service.price = price
    service.category = category
    db.commit()
    log_audit_action(db, current_user.id, "service_updated", "service", service.id, ip_address=request.client.host)
    return RedirectResponse("/services", status_code=303)

@app.get("/services/{service_id}/delete")
def delete_service(service_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    service = db.query(Service).filter(Service.id == service_id).first()
    if service:
        service.is_active = False
        db.commit()
        log_audit_action(db, current_user.id, "service_deleted", "service", service.id, service.name, request.client.host)
        request.session["flash"] = "Service removed."
    return RedirectResponse("/services", status_code=303)

# =========================
# INITIALIZATION
# =========================

def create_default_admin():
    db = SessionLocal()
    try:
        existing = db.query(User).filter(User.role == UserRole.ADMIN).first()
        if not existing:
            username = os.getenv("ADMIN_USER")
            password = os.getenv("ADMIN_PASS")
            if username and password:
                admin = User(username=username, password=hash_password(password), role=UserRole.ADMIN)
                db.add(admin)
                db.commit()
                print(f"Admin created: {username}")
            else:
                print("Set ADMIN_USER and ADMIN_PASS in .env")
    except Exception as e:
        print(f"Error: {e}")
    finally:
        db.close()


@app.on_event("startup")
async def startup_event():
    create_default_admin()
   

    # =========================
# USER MANAGEMENT (Admin Only) - Additional Routes
# =========================

@app.get("/users", response_class=HTMLResponse)
def list_users(request: Request, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """View all users - Admin only"""
    users = db.query(User).order_by(User.created_at.desc()).all()
    return templates.TemplateResponse("users.html", {
        "request": request,
        "users": users,
        "current_user": current_user
    })

@app.get("/users/{user_id}/edit", response_class=HTMLResponse)
def edit_user_page(user_id: int, request: Request, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Edit user form - Admin only"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        request.session["flash"] = "User not found."
        return RedirectResponse("/users", status_code=303)
    
    return templates.TemplateResponse("edit_user.html", {
        "request": request,
        "user": user,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/users/{user_id}/edit")
def update_user(
    user_id: int,
    request: Request,
    role: str = Form(...),
    is_active: str = Form(None),  # Checkbox returns None if unchecked
    new_password: str = Form(None),  # Optional password reset
    csrf_token: str = Form(...),
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Update user details - Admin only"""
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        request.session["flash"] = "User not found."
        return RedirectResponse("/users", status_code=303)
    
    # Prevent deactivating yourself
    if user.id == current_user.id and is_active is None:
        request.session["flash"] = "You cannot deactivate your own account."
        return RedirectResponse(f"/users/{user_id}/edit", status_code=303)
    
    # Update role
    user.role = UserRole.ADMIN if role == "admin" else UserRole.USER
    
    # Update active status
    user.is_active = is_active == "on"
    
    # Update password if provided
    if new_password and len(new_password) >= 8:
        user.password = hash_password(new_password)
        log_audit_action(db, current_user.id, "user_password_reset", "user", user.id, 
                        f"Password reset for {user.username}", request.client.host)
    
    db.commit()
    
    log_audit_action(db, current_user.id, "user_updated", "user", user.id, 
                    f"Updated {user.username} - Role: {role}, Active: {user.is_active}", request.client.host)
    
    request.session["flash"] = f"User '{user.username}' updated successfully."
    return RedirectResponse("/users", status_code=303)

@app.get("/users/{user_id}/delete")
def delete_user(user_id: int, request: Request, current_user: User = Depends(require_admin), db: Session = Depends(get_db)):
    """Delete user - Admin only with safety checks"""
    user = db.query(User).filter(User.id == user_id).first()
    
    if not user:
        request.session["flash"] = "User not found."
        return RedirectResponse("/users", status_code=303)
    
    # Prevent self-deletion
    if user.id == current_user.id:
        request.session["flash"] = "You cannot delete your own account."
        return RedirectResponse("/users", status_code=303)
    
    # Prevent deleting last admin
    if user.role == UserRole.ADMIN:
        admin_count = db.query(User).filter(User.role == UserRole.ADMIN, User.is_active == True).count()
        if admin_count <= 1:
            request.session["flash"] = "Cannot delete the last active administrator."
            return RedirectResponse("/users", status_code=303)
    
    # Check if user has created any clients, quotes, or invoices
    clients_count = db.query(Client).filter(Client.created_by_id == user.id).count()
    quotes_count = db.query(Quote).filter(Quote.created_by_id == user.id).count()
    invoices_count = db.query(Invoice).filter(Invoice.created_by_id == user.id).count()
    
    if clients_count > 0 or quotes_count > 0 or invoices_count > 0:
        # Soft delete - just deactivate instead of hard delete
        user.is_active = False
        db.commit()
        request.session["flash"] = f"User '{user.username}' has been deactivated (has existing records)."
        log_audit_action(db, current_user.id, "user_deactivated", "user", user.id, 
                        f"Deactivated {user.username} (has records)", request.client.host)
    else:
        # Hard delete for users with no records
        username = user.username
        db.delete(user)
        db.commit()
        request.session["flash"] = f"User '{username}' has been permanently deleted."
        log_audit_action(db, current_user.id, "user_deleted", "user", user_id, 
                        f"Deleted {username}", request.client.host)
    
    return RedirectResponse("/users", status_code=303)

@app.get("/clients/{client_id}/export")
def export_client_report(client_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Export client history to Excel"""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get data
    if current_user.role == UserRole.ADMIN:
        quotes = db.query(Quote).filter(Quote.client_id == client_id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id).all()
    else:
        quotes = db.query(Quote).filter(Quote.client_id == client_id, Quote.created_by_id == current_user.id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id, Invoice.created_by_id == current_user.id).all()
    
    # Create Excel workbook
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="49BEF5", end_color="49BEF5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    money_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Client Info
    ws_info = wb.active
    ws_info.title = "Client Info"
    
    ws_info['A1'] = "Client Report"
    ws_info['A1'].font = Font(bold=True, size=16)
    ws_info.merge_cells('A1:B1')
    
    info_data = [
        ("Client Name", client.name),
        ("Client Code", client.client_code),
        ("Email", client.email or "N/A"),
        ("Phone", client.phone or "N/A"),
        ("Address", client.address or "N/A"),
        ("Billing Name", client.billing_name or "N/A"),
        ("Billing Email", client.billing_email or "N/A"),
        ("VAT Number", client.vat_number or "N/A"),
        ("Tax Number", client.tax_number or "N/A"),
        ("Payment Terms", client.payment_terms or "N/A"),
        ("Created", client.created_at.strftime("%Y-%m-%d") if client.created_at else "N/A")
    ]
    
    for idx, (label, value) in enumerate(info_data, start=3):
        ws_info[f'A{idx}'] = label
        ws_info[f'A{idx}'].font = Font(bold=True)
        ws_info[f'B{idx}'] = value
    
    # Adjust column widths
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 40
    
    # Sheet 2: Quotes
    ws_quotes = wb.create_sheet("Quotes")
    headers = ["Quote #", "Date", "Total (R)", "Status", "Converted"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_quotes.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for idx, quote in enumerate(quotes, start=2):
        ws_quotes.cell(idx, 1, f"Q-{quote.quote_number:04d}").border = border
        ws_quotes.cell(idx, 2, quote.created_at.strftime("%Y-%m-%d")).border = border
        ws_quotes.cell(idx, 3, quote.total).border = border
        ws_quotes.cell(idx, 3).font = money_font
        ws_quotes.cell(idx, 3).number_format = '"R" #,##0.00'
        ws_quotes.cell(idx, 4, quote.status).border = border
        ws_quotes.cell(idx, 5, "Yes" if quote.converted else "No").border = border
    
    # Adjust column widths for quotes
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_quotes.column_dimensions[col].width = 15
    
    # Sheet 3: Invoices
    ws_inv = wb.create_sheet("Invoices")
    headers = ["Invoice #", "Date", "Total (R)", "Status", "Paid Date"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_inv.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    for idx, inv in enumerate(invoices, start=2):
        inv_num = f"{client.client_code[:3].upper()}-INV-{inv.invoice_number:04d}"
        ws_inv.cell(idx, 1, inv_num).border = border
        ws_inv.cell(idx, 2, inv.created_at.strftime("%Y-%m-%d")).border = border
        ws_inv.cell(idx, 3, inv.total).border = border
        ws_inv.cell(idx, 3).font = money_font
        ws_inv.cell(idx, 3).number_format = '"R" #,##0.00'
        ws_inv.cell(idx, 4, "Paid" if inv.paid else "Unpaid").border = border
        paid_date = inv.paid_at.strftime("%Y-%m-%d") if inv.paid and inv.paid_at else "N/A"
        ws_inv.cell(idx, 5, paid_date).border = border
    
    # Adjust column widths for invoices
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_inv.column_dimensions[col].width = 15
    
    # Sheet 4: Summary
    ws_summary = wb.create_sheet("Summary")
    
    total_quotes = len(quotes)
    total_invoices = len(invoices)
    total_paid = sum(inv.total for inv in invoices if inv.paid)
    total_unpaid = sum(inv.total for inv in invoices if not inv.paid)
    total_quotes_value = sum(q.total for q in quotes)
    
    summary_data = [
        ("Summary Statistics", ""),
        ("", ""),
        ("Total Quotes", total_quotes),
        ("Total Quotes Value", f"R {total_quotes_value:,.2f}"),
        ("", ""),
        ("Total Invoices", total_invoices),
        ("Total Paid", f"R {total_paid:,.2f}"),
        ("Total Unpaid", f"R {total_unpaid:,.2f}"),
        ("", ""),
        ("Conversion Rate", f"{(len([q for q in quotes if q.converted])/total_quotes*100 if total_quotes else 0):.1f}%")
    ]
    
    for idx, (label, value) in enumerate(summary_data, start=1):
        if label == "Summary Statistics":
            ws_summary.cell(idx, 1, label).font = title_font
            ws_summary.merge_cells(f'A{idx}:B{idx}')
        else:
            ws_summary.cell(idx, 1, label).font = Font(bold=True) if label else Font()
            ws_summary.cell(idx, 2, value)
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    
    # Save to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Client_Report_{client.client_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/clients/{client_id}/export/excel")
def export_client_excel(client_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Export client history to branded Excel"""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user.role == UserRole.ADMIN:
        quotes = db.query(Quote).filter(Quote.client_id == client_id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id).all()
    else:
        quotes = db.query(Quote).filter(Quote.client_id == client_id, Quote.created_by_id == current_user.id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id, Invoice.created_by_id == current_user.id).all()
    
    wb = Workbook()
    
    # Styles
    brand_color = "49BEF5"
    header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="2C3E50")
    subtitle_font = Font(bold=True, size=12, color="2C3E50")
    money_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin', color="CCCCCC"),
        right=Side(style='thin', color="CCCCCC"),
        top=Side(style='thin', color="CCCCCC"),
        bottom=Side(style='thin', color="CCCCCC")
    )
    
    # Sheet 1: Cover Page
    ws_cover = wb.active
    ws_cover.title = "Report Cover"
    
    # Try to add logo
    logo_path = "app/static/logo.png"
    if os.path.exists(logo_path):
        img = XLImage(logo_path)
        img.width = 300
        img.height = 100
        ws_cover.add_image(img, 'B2')
    
    ws_cover['B8'] = "CLIENT REPORT"
    ws_cover['B8'].font = Font(bold=True, size=24, color=brand_color)
    ws_cover['B9'] = client.name
    ws_cover['B9'].font = Font(bold=True, size=18)
    ws_cover['B10'] = f"Client Code: {client.client_code}"
    ws_cover['B10'].font = Font(size=12, color="666666")
    ws_cover['B11'] = f"Generated: {datetime.now().strftime('%d %B %Y')}"
    
    # Company footer on cover
    ws_cover['B20'] = "Umvuzo Media (Pty) Ltd"
    ws_cover['B20'].font = Font(bold=True, size=12)
    ws_cover['B21'] = "4 Veldblom Street, Terenure, Kempton Park, 1619"
    ws_cover['B22'] = "Tel: +27 61 213 0052 | info@umvuzomedia.co.za | www.umvuzomedia.co.za"
    ws_cover['B23'] = "Reg: 2012/137462/07"
    
    for row in range(20, 24):
        ws_cover[f'B{row}'].font = Font(size=10, color="666666")
    
    # Sheet 2: Client Details
    ws_info = wb.create_sheet("Client Details")
    
    ws_info['A1'] = "CLIENT INFORMATION"
    ws_info['A1'].font = title_font
    ws_info.merge_cells('A1:C1')
    
    info_rows = [
        ("Company Name", client.name),
        ("Client Code", client.client_code),
        ("Email Address", client.email or "N/A"),
        ("Phone Number", client.phone or "N/A"),
        ("Physical Address", client.address or "N/A"),
        ("", ""),
        ("BILLING INFORMATION", ""),
        ("Billing Name", client.billing_name or client.name),
        ("Billing Email", client.billing_email or client.email or "N/A"),
        ("Billing Address", client.billing_address or client.address or "N/A"),
        ("VAT Number", client.vat_number or "N/A"),
        ("Tax Number", client.tax_number or "N/A"),
        ("Payment Terms", client.payment_terms or "30 Days")
    ]
    
    for idx, (label, value) in enumerate(info_rows, start=3):
        if label == "BILLING INFORMATION":
            ws_info[f'A{idx}'] = label
            ws_info[f'A{idx}'].font = subtitle_font
            ws_info.merge_cells(f'A{idx}:C{idx}')
        else:
            ws_info[f'A{idx}'] = label
            ws_info[f'A{idx}'].font = Font(bold=True)
            ws_info[f'B{idx}'] = value
            ws_info.merge_cells(f'B{idx}:C{idx}')
    
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 35
    ws_info.column_dimensions['C'].width = 20
    
    # Sheet 3: Quotes
    ws_quotes = wb.create_sheet("Quotes History")
    
    ws_quotes['A1'] = "QUOTATION HISTORY"
    ws_quotes['A1'].font = title_font
    ws_quotes.merge_cells('A1:E1')
    
    headers = ["Quote Number", "Date Created", "Amount (R)", "Status", "Converted to Invoice"]
    for col, header in enumerate(headers, 1):
        cell = ws_quotes.cell(3, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for idx, quote in enumerate(quotes, start=4):
        ws_quotes.cell(idx, 1, f"Q-{quote.quote_number:04d}").border = border
        ws_quotes.cell(idx, 1).alignment = Alignment(horizontal='center')
        
        ws_quotes.cell(idx, 2, quote.created_at.strftime("%d %b %Y")).border = border
        ws_quotes.cell(idx, 2).alignment = Alignment(horizontal='center')
        
        ws_quotes.cell(idx, 3, quote.total).border = border
        ws_quotes.cell(idx, 3).font = money_font
        ws_quotes.cell(idx, 3).number_format = '"R" #,##0.00'
        ws_quotes.cell(idx, 3).alignment = Alignment(horizontal='right')
        
        ws_quotes.cell(idx, 4, quote.status).border = border
        ws_quotes.cell(idx, 4).alignment = Alignment(horizontal='center')
        
        ws_quotes.cell(idx, 5, "Yes ✓" if quote.converted else "No").border = border
        ws_quotes.cell(idx, 5).alignment = Alignment(horizontal='center')
        if quote.converted:
            ws_quotes.cell(idx, 5).font = Font(color="28a745", bold=True)
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_quotes.column_dimensions[col].width = 18
    
    # Sheet 4: Invoices
    ws_inv = wb.create_sheet("Invoice History")
    
    ws_inv['A1'] = "INVOICE HISTORY"
    ws_inv['A1'].font = title_font
    ws_inv.merge_cells('A1:E1')
    
    headers = ["Invoice Number", "Date Created", "Amount (R)", "Status", "Paid Date"]
    for col, header in enumerate(headers, 1):
        cell = ws_inv.cell(3, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    for idx, inv in enumerate(invoices, start=4):
        inv_num = f"{client.client_code[:3].upper()}-INV-{inv.invoice_number:04d}"
        ws_inv.cell(idx, 1, inv_num).border = border
        ws_inv.cell(idx, 1).alignment = Alignment(horizontal='center')
        
        ws_inv.cell(idx, 2, inv.created_at.strftime("%d %b %Y")).border = border
        ws_inv.cell(idx, 2).alignment = Alignment(horizontal='center')
        
        ws_inv.cell(idx, 3, inv.total).border = border
        ws_inv.cell(idx, 3).font = money_font
        ws_inv.cell(idx, 3).number_format = '"R" #,##0.00'
        ws_inv.cell(idx, 3).alignment = Alignment(horizontal='right')
        
        status_cell = ws_inv.cell(idx, 4, "PAID" if inv.paid else "UNPAID")
        status_cell.border = border
        status_cell.alignment = Alignment(horizontal='center')
        status_cell.font = Font(color="28a745" if inv.paid else "dc3545", bold=True)
        
        paid_date = inv.paid_at.strftime("%d %b %Y") if inv.paid and inv.paid_at else "-"
        ws_inv.cell(idx, 5, paid_date).border = border
        ws_inv.cell(idx, 5).alignment = Alignment(horizontal='center')
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_inv.column_dimensions[col].width = 18
    
    # Sheet 5: Financial Summary
    ws_sum = wb.create_sheet("Financial Summary")
    
    ws_sum['A1'] = "FINANCIAL SUMMARY"
    ws_sum['A1'].font = title_font
    ws_sum.merge_cells('A1:C1')
    
    total_quotes = len(quotes)
    total_quotes_val = sum(q.total for q in quotes)
    total_invoices = len(invoices)
    total_paid = sum(inv.total for inv in invoices if inv.paid)
    total_unpaid = sum(inv.total for inv in invoices if not inv.paid)
    conversion_rate = (len([q for q in quotes if q.converted])/total_quotes*100) if total_quotes else 0
    
    summary_data = [
        ("Metric", "Value", "Notes"),
        ("Total Quotes Issued", total_quotes, "Lifetime count"),
        ("Total Quotes Value", f"R {total_quotes_val:,.2f}", "Potential revenue"),
        ("", "", ""),
        ("Total Invoices Generated", total_invoices, "Converted quotes"),
        ("Total Paid", f"R {total_paid:,.2f}", "Revenue received"),
        ("Total Outstanding", f"R {total_unpaid:,.2f}", "Revenue pending"),
        ("", "", ""),
        ("Quote-to-Invoice Conversion", f"{conversion_rate:.1f}%", "Success rate"),
        ("Average Quote Value", f"R {(total_quotes_val/total_quotes if total_quotes else 0):,.2f}", "Per quote"),
        ("Average Invoice Value", f"R {(total_paid/(len([i for i in invoices if i.paid]) or 1)):,.2f}", "Paid invoices only")
    ]
    
    for idx, (metric, value, notes) in enumerate(summary_data, start=3):
        if idx == 3:  # Header row
            ws_sum.cell(idx, 1, metric).fill = header_fill
            ws_sum.cell(idx, 1).font = header_font
            ws_sum.cell(idx, 2, value).fill = header_fill
            ws_sum.cell(idx, 2).font = header_font
            ws_sum.cell(idx, 3, notes).fill = header_fill
            ws_sum.cell(idx, 3).font = header_font
        else:
            ws_sum.cell(idx, 1, metric).font = Font(bold=True) if metric else Font()
            ws_sum.cell(idx, 2, value).font = money_font if isinstance(value, str) and value.startswith("R") else Font()
            ws_sum.cell(idx, 3, notes).font = Font(italic=True, color="666666")
        
        for col in [1, 2, 3]:
            ws_sum.cell(idx, col).border = border
            ws_sum.cell(idx, col).alignment = Alignment(vertical='center')
    
    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 20
    ws_sum.column_dimensions['C'].width = 25
    
    # Add footer to all sheets
    for ws in [ws_cover, ws_info, ws_quotes, ws_inv, ws_sum]:
        ws.oddFooter.center.text = "Umvuzo Media (Pty) Ltd | Reg: 2012/137462/07 | info@umvuzomedia.co.za"
        ws.oddFooter.center.size = 9
        ws.oddFooter.center.color = "666666"
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Client_Report_{client.client_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/clients/{client_id}/export/pdf")
def export_client_pdf(client_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    """Export client history to branded PDF report"""
    client = db.get(Client, client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    if current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if current_user.role == UserRole.ADMIN:
        quotes = db.query(Quote).filter(Quote.client_id == client_id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id).all()
    else:
        quotes = db.query(Quote).filter(Quote.client_id == client_id, Quote.created_by_id == current_user.id).all()
        invoices = db.query(Invoice).filter(Invoice.client_id == client_id, Invoice.created_by_id == current_user.id).all()
    
    # Setup PDF
    filename = f"Client_Report_{client.client_code}_{datetime.now().strftime('%Y%m%d')}.pdf"
    doc = SimpleDocTemplate(
        filename,
        pagesize=A4,
        rightMargin=0.6*inch,
        leftMargin=0.6*inch,
        topMargin=0.8*inch,
        bottomMargin=0.8*inch
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Colors
    brand_color = colors.HexColor("#49BEF5")
    dark_blue = colors.HexColor("#2C3E50")
    light_gray = colors.HexColor("#F8F9FA")
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=dark_blue,
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Header style
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=brand_color,
        spaceAfter=12,
        fontName='Helvetica-Bold'
    )
    
    # Normal style
    normal_style = styles["Normal"]
    normal_style.fontSize = 10
    
    # Logo
    logo_path = "app/static/logo.png"
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=2*inch, height=1.2*inch)
        elements.append(logo)
        elements.append(Spacer(1, 0.2*inch))
    
    # Report Title
    elements.append(Paragraph("CLIENT REPORT", title_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Client Name
    elements.append(Paragraph(f"<b>{client.name}</b>", ParagraphStyle(
        'ClientName',
        parent=styles['Normal'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=6
    )))
    elements.append(Paragraph(f"Client Code: {client.client_code} | Generated: {datetime.now().strftime('%d %B %Y')}", ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#666666"),
        spaceAfter=20
    )))
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Client Info Section
    elements.append(Paragraph("CLIENT INFORMATION", header_style))
    
    info_data = [
        ["Company Name:", client.name],
        ["Email:", client.email or "N/A"],
        ["Phone:", client.phone or "N/A"],
        ["Address:", client.address or "N/A"],
        ["Billing Name:", client.billing_name or client.name],
        ["Billing Email:", client.billing_email or client.email or "N/A"],
        ["VAT Number:", client.vat_number or "N/A"],
        ["Payment Terms:", client.payment_terms or "30 Days"]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, -1), light_gray),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#DDDDDD")),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Quotes Section
    if quotes:
        elements.append(Paragraph("QUOTATION HISTORY", header_style))
        
        quote_data = [["Quote #", "Date", "Amount", "Status", "Converted"]]
        for q in quotes:
            quote_data.append([
                f"Q-{q.quote_number:04d}",
                q.created_at.strftime("%d %b %Y"),
                f"R {q.total:,.2f}",
                q.status,
                "Yes" if q.converted else "No"
            ])
        
        quote_table = Table(quote_data, colWidths=[1*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
        quote_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, 1), (4, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
        ]))
        elements.append(quote_table)
        elements.append(Spacer(1, 0.2*inch))
    
    # Invoices Section
    if invoices:
        if quotes:
            elements.append(Spacer(1, 0.2*inch))
        
        elements.append(Paragraph("INVOICE HISTORY", header_style))
        
        inv_data = [["Invoice #", "Date", "Amount", "Status", "Paid Date"]]
        for inv in invoices:
            inv_num = f"{client.client_code[:3].upper()}-INV-{inv.invoice_number:04d}"
            paid_date = inv.paid_at.strftime("%d %b %Y") if inv.paid and inv.paid_at else "-"
            status_color = "green" if inv.paid else "red"
            
            inv_data.append([
                inv_num,
                inv.created_at.strftime("%d %b %Y"),
                f"R {inv.total:,.2f}",
                "PAID" if inv.paid else "UNPAID",
                paid_date
            ])
        
        inv_table = Table(inv_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
        inv_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#28a745")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, 1), (4, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 8),
            ('TEXTCOLOR', (3, 1), (3, -1), colors.HexColor("#28a745"), lambda x, y, z: z[3] == "PAID"),
            ('TEXTCOLOR', (3, 1), (3, -1), colors.HexColor("#dc3545"), lambda x, y, z: z[3] == "UNPAID"),
        ]))
        elements.append(inv_table)
    
    # Financial Summary
    elements.append(Spacer(1, 0.4*inch))
    elements.append(Paragraph("FINANCIAL SUMMARY", header_style))
    
    total_quotes = len(quotes)
    total_quotes_val = sum(q.total for q in quotes)
    total_invoices = len(invoices)
    total_paid = sum(inv.total for inv in invoices if inv.paid)
    total_unpaid = sum(inv.total for inv in invoices if not inv.paid)
    
    summary_data = [
        ["Total Quotes:", str(total_quotes), f"R {total_quotes_val:,.2f}"],
        ["Total Invoices:", str(total_invoices), f"R {total_paid + total_unpaid:,.2f}"],
        ["Total Paid:", "", f"R {total_paid:,.2f}"],
        ["Total Outstanding:", "", f"R {total_unpaid:,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LINEABOVE', (0, 0), (-1, 0), 1.5, brand_color),
        ('LINEBELOW', (0, -1), (-1, -1), 1.5, brand_color),
    ]))
    elements.append(summary_table)
    
    # Footer with company info
    elements.append(Spacer(1, 0.5*inch))
    footer_text = """
    <para alignment="center" fontSize="9" textColor="#666666">
    <b>Umvuzo Media (Pty) Ltd</b><br/>
    4 Veldblom Street, Terenure, Kempton Park, 1619<br/>
    Tel: +27 61 213 0052 | info@umvuzomedia.co.za | www.umvuzomedia.co.za<br/>
    Reg: 2012/137462/07
    </para>
    """
    elements.append(Paragraph(footer_text, styles["Normal"]))
    
    doc.build(elements)
    
    return FileResponse(filename, media_type="application/pdf", filename=filename)


@app.get("/pricing", response_class=HTMLResponse)
def pricing_page(request: Request):
    return templates.TemplateResponse("pricing.html", {"request": request})

# =========================
# PREVIEW ROUTES (HTML View)
# =========================

from datetime import timedelta

@app.get("/quotes/{quote_id}/preview", response_class=HTMLResponse)
def preview_quote(request: Request, quote_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client = db.get(Client, quote.client_id)
    items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).all()
    
    # Calculate valid until date
    valid_until = quote.created_at + timedelta(days=30)
    
    return templates.TemplateResponse("quote_preview.html", {
        "request": request,
        "quote": quote,
        "client": client,
        "items": items,
        "current_user": current_user,
        "valid_until": valid_until
    })

@app.get("/invoices/{invoice_id}/preview", response_class=HTMLResponse)
def preview_invoice(request: Request, invoice_id: int, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    if current_user.role != UserRole.ADMIN and invoice.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    client = db.get(Client, invoice.client_id)
    items = db.query(InvoiceItem).filter(InvoiceItem.invoice_id == invoice.id).all()
    
    # Calculate due date
    due_date = invoice.created_at + timedelta(days=30)
    
    return templates.TemplateResponse("invoice_preview.html", {
        "request": request,
        "invoice": invoice,
        "client": client,
        "items": items,
        "current_user": current_user,
        "due_date": due_date
    })


    # =========================
# EDIT QUOTE
# =========================

@app.get("/quotes/{quote_id}/edit", response_class=HTMLResponse)
def edit_quote_page(quote_id: int, request: Request, current_user: User = Depends(require_auth), db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if quote.converted:
        request.session["flash"] = "Cannot edit - quote already converted to invoice."
        return RedirectResponse("/quotes-page", status_code=303)
    
    client = db.get(Client, quote.client_id)
    items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).all()
    services = db.query(Service).filter(Service.is_active == True).order_by(Service.name).all()
    clients = db.query(Client).all()
    
    return templates.TemplateResponse("edit_quote.html", {
        "request": request,
        "quote": quote,
        "client": client,
        "items": items,
        "services": services,
        "clients": clients,
        "csrf_token": request.session.get("csrf_token"),
        "current_user": current_user
    })

@app.post("/quotes/{quote_id}/edit")
def update_quote(
    quote_id: int,
    request: Request,
    client_id: int = Form(...),
    items_data: str = Form(...),
    csrf_token: str = Form(...),
    current_user: User = Depends(require_auth),
    db: Session = Depends(get_db)
):
    if csrf_token != request.session.get("csrf_token"):
        raise HTTPException(status_code=403, detail="Invalid CSRF token")
    
    quote = db.query(Quote).filter(Quote.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")
    
    if current_user.role != UserRole.ADMIN and quote.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Access denied")
    
    if quote.converted:
        request.session["flash"] = "Cannot edit converted quote."
        return RedirectResponse("/quotes-page", status_code=303)
    
    # Verify client access
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client or (current_user.role != UserRole.ADMIN and client.created_by_id != current_user.id):
        raise HTTPException(status_code=403, detail="Invalid client")
    
    try:
        items = json.loads(items_data)
        if not items:
            request.session["flash"] = "Add at least one item."
            return RedirectResponse(f"/quotes/{quote_id}/edit", status_code=303)
    except:
        request.session["flash"] = "Invalid data."
        return RedirectResponse(f"/quotes/{quote_id}/edit", status_code=303)
    
    try:
        # Update client if changed
        quote.client_id = client_id
        
        # Delete old items
        db.query(QuoteItem).filter(QuoteItem.quote_id == quote.id).delete()
        
        # Add new items
        total = 0
        for item in items:
            line = float(item["unit_cost"]) * float(item["quantity"])
            total += line
            db.add(QuoteItem(
                quote_id=quote.id,
                description=item["description"],
                unit_cost=float(item["unit_cost"]),
                quantity=float(item["quantity"])
            ))
        
        quote.total = total
        db.commit()
        
        log_audit_action(db, current_user.id, "quote_updated", "quote", quote.id, 
                        f"Updated Q-{quote.quote_number:04d}", request.client.host)
        
        request.session["flash"] = f"Quote Q-{quote.quote_number:04d} updated successfully!"
        return RedirectResponse("/quotes-page", status_code=303)
        
    except Exception as e:
        db.rollback()
        request.session["flash"] = "Error updating quote."
        return RedirectResponse(f"/quotes/{quote_id}/edit", status_code=303)